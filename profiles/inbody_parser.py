"""
Parser de exports InBody en CSV o Excel.

La app InBody exporta una fila por test con columnas en espanol.
Mapeamos solo los campos que existen en el modelo InbodyReport;
columnas extra se ignoran. Valores '-' o vacios se tratan como None.
"""
import csv
import unicodedata
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


# Columna del CSV/Excel (normalizada) -> campo del modelo InbodyReport.
# La normalizacion baja a minusculas, quita acentos y colapsa espacios,
# asi toleramos pequenas variaciones del export.
COLUMNA_A_CAMPO = {
    'peso(kg)':                          'peso',
    'percentaje de grasa corporal(%)':   'porcentaje_grasa',   # typo del export real
    'porcentaje de grasa corporal(%)':   'porcentaje_grasa',   # por si lo corrigen
    'masa muscular(kg)':                 'masa_muscular',
    'grasa corporal(kg)':                'masa_grasa_kg',
    'imc(kg/m2)':                        'imc',
    'puntuacion inbody':                 'puntuacion_inbody',
    'tasa metabolica basal(kcal)':       'tasa_metabolica_basal',
    'agua corporal total(l)':            'agua_corporal',
}

CAMPOS_ENTEROS = {'puntuacion_inbody', 'tasa_metabolica_basal'}


def _normalize_header(s) -> str:
    if s is None:
        return ''
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return ' '.join(s.lower().split())


def _to_decimal(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s == '-':
        return None
    try:
        return float(s.replace(',', '.'))
    except ValueError:
        return None


def _to_int(raw):
    val = _to_decimal(raw)
    return int(val) if val is not None else None


def _parse_fecha(raw):
    """'20260511091108' -> '2026-05-11'. None si no se puede parsear."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s == '-':
        return None
    if s.isdigit() and len(s) >= 8:
        try:
            return date(int(s[0:4]), int(s[4:6]), int(s[6:8])).isoformat()
        except ValueError:
            pass
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _leer_csv(path: Path):
    for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
        try:
            with open(path, newline='', encoding=encoding) as f:
                reader = csv.reader(f)
                headers = next(reader, None)
                if not headers:
                    return None, None
                fila = next(reader, None)
                if not fila:
                    return None, None
                return headers, fila
        except UnicodeDecodeError:
            continue
    return None, None


def _leer_excel(path: Path):
    if not load_workbook:
        return None, None
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return None, None
        fila = next(rows, None)
        if not fila:
            return None, None
        return list(headers), list(fila)
    except Exception:
        return None, None


def parse_inbody(file_path) -> dict:
    """
    Extrae metricas del export InBody (CSV o Excel).
    Devuelve un dict con los campos detectados; los demas no aparecen.
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == '.csv':
        headers, fila = _leer_csv(path)
    elif suffix in ('.xlsx', '.xls'):
        headers, fila = _leer_excel(path)
    else:
        return {}

    if not headers or not fila:
        return {}

    resultado = {}
    for header, valor in zip(headers, fila):
        norm = _normalize_header(header)

        if norm == 'fecha':
            fecha = _parse_fecha(valor)
            if fecha:
                resultado['fecha_test'] = fecha
            continue

        campo = COLUMNA_A_CAMPO.get(norm)
        if not campo:
            continue

        val = _to_int(valor) if campo in CAMPOS_ENTEROS else _to_decimal(valor)
        if val is not None:
            resultado[campo] = val

    return resultado
